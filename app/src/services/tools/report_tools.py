"""Report generation tools — Excel report builder.

`generate_excel_report` creates styled .xlsx files from financial analysis
data. The agent calls this after `financial_analyzer` to provide a
downloadable report.
"""
from __future__ import annotations

import json
import logging
import os
import re
import uuid
from datetime import datetime

from langchain_core.tools import tool

logger = logging.getLogger(__name__)


def _get_reports_dir() -> str:
    from src.core.config import get_settings
    d = get_settings().reports_dir
    os.makedirs(d, exist_ok=True)
    return d


@tool
def generate_excel_report(analysis_text: str) -> str:
    """Generate a downloadable Excel report from financial analysis results.

    Input `analysis_text` should be the output from the financial_analyzer tool
    (contains markdown + embedded JSON data). Or, it can be a JSON string with
    financial data directly.

    The Excel file will contain:
    - Sheet 1: Input Data (key financial figures)
    - Sheet 2: Rasio Keuangan (computed ratios with color-coded health status)
    - Sheet 3: Investment Assessment (score + recommendation)

    Returns the download URL path for the generated report.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "Error: openpyxl is not installed. Cannot generate Excel reports."

    # Extract JSON data from the analysis text
    data = None

    # Try to find embedded financial data in HTML comment
    match = re.search(r"<!-- FINANCIAL_DATA:(.*?) -->", analysis_text)
    if match:
        try:
            data = json.loads(match.group(1))
        except json.JSONDecodeError:
            pass

    # Try parsing the entire input as JSON
    if data is None:
        try:
            data = json.loads(analysis_text)
        except json.JSONDecodeError:
            pass

    if data is None:
        return (
            "Error: Could not parse financial data. Please run financial_analyzer first "
            "and pass its output to this tool."
        )

    # Create workbook
    wb = Workbook()

    # -- Styles --
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="2C3E50")
    subtitle_font = Font(name="Calibri", bold=True, size=11, color="34495E")
    normal_font = Font(name="Calibri", size=10)
    number_font = Font(name="Calibri", size=10)
    green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    def _style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

    def _style_cell(ws, row, col, align=left_align):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = align
        return cell

    # ===== Sheet 1: Input Data =====
    ws1 = wb.active
    ws1.title = "Data Keuangan"
    ws1.sheet_properties.tabColor = "2C3E50"

    ws1.cell(row=1, column=1, value="📊 Data Keuangan Input").font = title_font
    ws1.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color="95A5A6")

    input_data = data.get("input_data", {})
    label_map = {
        "revenue": "Revenue / Pendapatan",
        "cogs": "COGS / Harga Pokok Penjualan",
        "gross_profit": "Laba Kotor",
        "net_income": "Laba Bersih",
        "total_assets": "Total Aset",
        "total_equity": "Total Ekuitas",
        "total_debt": "Total Utang",
        "current_assets": "Aset Lancar",
        "current_liabilities": "Liabilitas Lancar",
        "inventory": "Persediaan",
        "prev_revenue": "Revenue Tahun Sebelumnya",
        "prev_net_income": "Laba Bersih Tahun Sebelumnya",
        "stock_price": "Harga Saham",
        "shares_outstanding": "Jumlah Saham Beredar",
    }

    row = 4
    ws1.cell(row=row, column=1, value="Item").font = header_font
    ws1.cell(row=row, column=2, value="Nilai").font = header_font
    _style_header(ws1, row, 2)

    for key, value in input_data.items():
        row += 1
        label = label_map.get(key, key.replace("_", " ").title())
        _style_cell(ws1, row, 1).value = label
        cell = _style_cell(ws1, row, 2, right_align)
        cell.value = value
        cell.number_format = '#,##0'

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 25

    # ===== Sheet 2: Rasio Keuangan =====
    ws2 = wb.create_sheet("Rasio Keuangan")
    ws2.sheet_properties.tabColor = "2980B9"

    ws2.cell(row=1, column=1, value="📈 Analisis Rasio Keuangan").font = title_font
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color="95A5A6")

    row = 4
    headers = ["Kategori", "Rasio", "Nilai", "Status"]
    for i, h in enumerate(headers, 1):
        ws2.cell(row=row, column=i, value=h)
    _style_header(ws2, row, len(headers))

    ratios = data.get("ratios", [])
    for r in ratios:
        row += 1
        _style_cell(ws2, row, 1).value = r.get("category", "")
        _style_cell(ws2, row, 2).value = r.get("name", "")
        _style_cell(ws2, row, 3, right_align).value = r.get("formatted", "")
        status_cell = _style_cell(ws2, row, 4, center_align)
        health = r.get("health", "")
        status_cell.value = health

        # Color code
        if "Sehat" in health:
            status_cell.fill = green_fill
        elif "Perhatian" in health:
            status_cell.fill = yellow_fill
        elif "Bahaya" in health:
            status_cell.fill = red_fill

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 22

    # ===== Sheet 3: Investment Assessment =====
    ws3 = wb.create_sheet("Assessment Investasi")
    ws3.sheet_properties.tabColor = "27AE60"

    ws3.cell(row=1, column=1, value="💡 Investment Assessment").font = title_font
    ws3.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color="95A5A6")

    row = 4
    score = data.get("score", 0)
    ws3.cell(row=row, column=1, value="Skor Fundamental").font = subtitle_font
    score_cell = ws3.cell(row=row, column=2, value=f"{score}/100")
    score_cell.font = Font(name="Calibri", bold=True, size=16)
    if score >= 70:
        score_cell.font = Font(name="Calibri", bold=True, size=16, color="27AE60")
    elif score >= 40:
        score_cell.font = Font(name="Calibri", bold=True, size=16, color="F39C12")
    else:
        score_cell.font = Font(name="Calibri", bold=True, size=16, color="E74C3C")

    row += 2
    stats = [
        ("Rasio Sehat", data.get("healthy", 0), green_fill),
        ("Perlu Perhatian", data.get("warning", 0), yellow_fill),
        ("Bahaya", data.get("danger", 0), red_fill),
        ("Total Rasio", data.get("total", 0), None),
    ]
    for label, val, fill in stats:
        ws3.cell(row=row, column=1, value=label).font = normal_font
        ws3.cell(row=row, column=1).border = thin_border
        cell = ws3.cell(row=row, column=2, value=val)
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.border = thin_border
        cell.alignment = center_align
        if fill:
            cell.fill = fill
        row += 1

    row += 1
    ws3.cell(row=row, column=1, value="Rekomendasi").font = subtitle_font
    row += 1
    rec = data.get("recommendation", "Tidak ada data")
    # Clean emoji for Excel
    clean_rec = re.sub(r"[🟢🟡🔴✅⚠️]", "", rec).strip()
    ws3.cell(row=row, column=1, value=clean_rec).font = Font(name="Calibri", size=11)
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 20

    # Save
    reports_dir = _get_reports_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    short_id = uuid.uuid4().hex[:6]
    filename = f"financial_report_{timestamp}_{short_id}.xlsx"
    filepath = os.path.join(reports_dir, filename)

    wb.save(filepath)
    logger.info(f"Excel report generated: {filepath}")

    return (
        f"✅ Excel report berhasil dibuat!\n"
        f"📥 Download: /v1/reports/{filename}\n"
        f"File: {filename}"
    )


__all__ = ["generate_excel_report"]
